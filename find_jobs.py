#!/usr/bin/env python3
"""
Salesforce Job Finder — Multi-Platform India
Platforms: LinkedIn, Indeed, Naukri, TimesJobs, Foundit, Shine,
           iimjobs, Glassdoor, Cutshort + Greenhouse & Lever APIs
"""
import asyncio
import re
import json
from datetime import datetime, date, timezone
from pathlib import Path
from playwright.async_api import async_playwright, Page
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_FILE = Path(__file__).parent / "salesforce_jobs.xlsx"
JOBS_JSON   = Path(__file__).parent / "jobs.json"

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# Broader keywords → more hits across all platforms
KEYWORDS = [
    "Salesforce Developer",
    "Salesforce Consultant",
    "Salesforce LWC",
    "Salesforce CPQ",
    "Salesforce Architect",
    "Salesforce Admin",
]

MAX_AGE_HOURS = 72   # 3 days


# ── helpers ────────────────────────────────────────────────────────────────────

def parse_age_hours(text: str) -> float:
    """Return job age in hours. 0 = just posted / unknown (keep it)."""
    text = (text or "").lower().strip()
    if not text:
        return 0.0

    # ISO datetime: 2026-07-24T10:30:00Z
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})", text)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                          int(m.group(4)), int(m.group(5)), tzinfo=timezone.utc)
            return max(0.0, (datetime.now(timezone.utc) - dt).total_seconds() / 3600)
        except Exception:
            pass

    # ISO date only: 2026-07-24
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})$", text)
    if m:
        try:
            d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return max(0.0, (date.today() - d).days * 24.0)
        except Exception:
            pass

    if any(t in text for t in ("just now", "moments ago", "less than")):
        return 0.5
    if "today" in text or "new today" in text or text == "posted today":
        return 4.0
    if "yesterday" in text or text == "posted yesterday":
        return 28.0
    # Workday "Posted X Days Ago" / "Posted X Months Ago"
    m = re.search(r"posted\s+(\d+)\s*day", text)
    if m: return int(m.group(1)) * 24.0
    m = re.search(r"posted\s+(\d+)\s*month", text)
    if m: return int(m.group(1)) * 720.0
    m = re.search(r"posted\s+(\d+)\s*week", text)
    if m: return int(m.group(1)) * 168.0

    m = re.search(r"(\d+)\s*sec", text)
    if m: return round(int(m.group(1)) / 3600, 2)
    m = re.search(r"(\d+)\s*min", text)
    if m: return round(int(m.group(1)) / 60, 2)
    m = re.search(r"(\d+)\s*(hour|hr)", text)
    if m: return float(m.group(1))
    m = re.search(r"(\d+)\s*day", text)
    if m: return int(m.group(1)) * 24.0
    m = re.search(r"(\d+)\s*week", text)
    if m: return int(m.group(1)) * 168.0
    m = re.search(r"(\d+)\s*month", text)
    if m: return int(m.group(1)) * 720.0

    return 0.0   # unknown → treat as fresh

def is_recent(age_h: float) -> bool:
    return age_h <= MAX_AGE_HOURS

def make_job(title, company, platform, url, age_txt, location="India", age_h=None):
    h = age_h if age_h is not None else parse_age_hours(age_txt)
    return {
        "title":     title.strip(),
        "company":   (company or "Unknown").strip(),
        "platform":  platform,
        "location":  (location or "India").strip(),
        "url":       url.strip(),
        "posted":    (age_txt or "Recent").strip(),
        "age_hours": round(h, 1),
        "age_days":  int(h / 24),
    }

async def safe_goto(page: Page, url: str, timeout=25000):
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        await page.wait_for_timeout(2500)
    except Exception:
        pass


# ── browser scrapers ───────────────────────────────────────────────────────────

async def scrape_linkedin(page: Page, keyword: str) -> list[dict]:
    jobs = []
    kw = keyword.replace(" ", "%20")
    # f_TPR=r259200 = past 3 days, geoId=102713980 = India
    await safe_goto(page, f"https://www.linkedin.com/jobs/search/?keywords={kw}&f_TPR=r259200&geoId=102713980&sortBy=DD")
    for card in await page.query_selector_all(".base-card, .job-search-card"):
        try:
            t_el = await card.query_selector(".base-search-card__title, h3")
            c_el = await card.query_selector(".base-search-card__subtitle, h4")
            a_el = await card.query_selector("a.base-card__full-link, a")
            d_el = await card.query_selector("time, .job-search-card__listdate")
            l_el = await card.query_selector(".job-search-card__location")
            if not (t_el and a_el): continue
            title   = (await t_el.inner_text()).strip()
            company = (await c_el.inner_text()).strip() if c_el else ""
            href    = await a_el.get_attribute("href") or ""
            loc     = (await l_el.inner_text()).strip() if l_el else "India"
            age_txt = ""
            if d_el:
                age_txt = await d_el.get_attribute("datetime") or await d_el.inner_text()
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "LinkedIn", href, age_txt, loc, h))
        except Exception:
            continue
    print(f"  LinkedIn '{keyword}': {len(jobs)}")
    return jobs


async def scrape_indeed(page: Page, keyword: str) -> list[dict]:
    jobs = []
    kw = keyword.replace(" ", "+")
    await safe_goto(page, f"https://in.indeed.com/jobs?q={kw}&l=India&sort=date&fromage=3")
    for sel in ["a.jcs-JobTitle", "[data-testid='job-title'] a", "h2.jobTitle a", ".job_seen_beacon h2 a"]:
        links = await page.query_selector_all(sel)
        if not links: continue
        for link in links:
            try:
                title = (await link.inner_text()).strip()
                href  = await link.get_attribute("href") or ""
                if not href.startswith("http"): href = "https://in.indeed.com" + href
                parent  = await link.evaluate_handle("el => el.closest('[data-jk], .job_seen_beacon, .resultContent')")
                company = age_txt = loc = ""
                try:
                    co = await parent.query_selector("[data-testid='company-name'], .companyName")
                    if co: company = (await co.inner_text()).strip()
                except Exception: pass
                try:
                    dt = await parent.query_selector("[data-testid='myJobsStateDate'], .date")
                    if dt: age_txt = (await dt.inner_text()).strip()
                except Exception: pass
                try:
                    lc = await parent.query_selector("[data-testid='text-location'], .companyLocation")
                    if lc: loc = (await lc.inner_text()).strip()
                except Exception: pass
                if title and href:
                    jobs.append(make_job(title, company, "Indeed", href, age_txt, loc or "India"))
            except Exception: continue
        if jobs: break
    print(f"  Indeed '{keyword}': {len(jobs)}")
    return jobs


async def scrape_naukri(page: Page, keyword: str) -> list[dict]:
    jobs = []
    kw = keyword.lower().replace(" ", "-")
    await safe_goto(page, f"https://www.naukri.com/{kw}-jobs-in-india?jobAge=3")
    await page.wait_for_timeout(2000)
    cards = await page.query_selector_all(
        "article.jobTuple, .srp-jobtuple-wrapper, .list .listContainer li"
    )
    for card in cards:
        try:
            t_el = await card.query_selector("a.title, .title a, a.row1")
            c_el = await card.query_selector(".comp-name, .companyInfo a")
            d_el = await card.query_selector(".job-post-day, span.fleft.grey-text")
            l_el = await card.query_selector(".loc, span.loc-link, .locWdth")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            loc     = (await l_el.inner_text()).strip() if l_el else "India"
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Naukri", href, age_txt, loc, h))
        except Exception: continue
    print(f"  Naukri '{keyword}': {len(jobs)}")
    return jobs


async def scrape_timesjobs(page: Page, keyword: str) -> list[dict]:
    jobs = []
    kw = keyword.replace(" ", "+")
    url = (
        f"https://www.timesjobs.com/candidate/job-search.html"
        f"?searchType=personalizedSearch&from=submit"
        f"&txtKeywords={kw}&txtLocation=India&postWeek=1"
    )
    await safe_goto(page, url)
    for card in await page.query_selector_all("li.clearfix.job-bx, .jobs-container li"):
        try:
            t_el = await card.query_selector("h2.heading-tit a, h2 a, .info-grp h2 a")
            c_el = await card.query_selector("h3.joblist-comp-name, .comp-info-detail h3")
            d_el = await card.query_selector(".dt-post, .job-time, span.sim-posted")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "TimesJobs", href, age_txt, "India", h))
        except Exception: continue
    print(f"  TimesJobs '{keyword}': {len(jobs)}")
    return jobs


async def scrape_foundit(page: Page, keyword: str) -> list[dict]:
    """Foundit.in — formerly Monster India"""
    jobs = []
    kw = keyword.replace(" ", "+")
    await safe_goto(page, f"https://www.foundit.in/srp/results?query={kw}&location=India&datePosted=3")
    await page.wait_for_timeout(2000)
    for card in await page.query_selector_all(".card-apply-content, .srpResultCardContainer, .jobCard"):
        try:
            t_el = await card.query_selector(".jobTitle a, .job-tittle a, h2 a")
            c_el = await card.query_selector(".companyName, .company-name")
            d_el = await card.query_selector(".posted-time, .job-date, .postedDate")
            l_el = await card.query_selector(".location, .job-location")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://www.foundit.in" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            loc     = (await l_el.inner_text()).strip() if l_el else "India"
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Foundit", href, age_txt, loc, h))
        except Exception: continue
    print(f"  Foundit '{keyword}': {len(jobs)}")
    return jobs


async def scrape_shine(page: Page, keyword: str) -> list[dict]:
    jobs = []
    kw = keyword.lower().replace(" ", "-")
    await safe_goto(page, f"https://www.shine.com/job-search/{kw}-jobs-in-india/")
    for card in await page.query_selector_all(".jobBox, .job-result-box, article"):
        try:
            t_el = await card.query_selector(".jobHeadline a, .job-title a, h2 a, h3 a")
            c_el = await card.query_selector(".companyName a, .company-name")
            d_el = await card.query_selector(".postedDate, .job-date, .date")
            l_el = await card.query_selector(".location, .city")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://www.shine.com" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            loc     = (await l_el.inner_text()).strip() if l_el else "India"
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Shine", href, age_txt, loc, h))
        except Exception: continue
    print(f"  Shine '{keyword}': {len(jobs)}")
    return jobs


async def scrape_iimjobs(page: Page, keyword: str) -> list[dict]:
    """iimjobs.com — premium India jobs"""
    jobs = []
    kw = keyword.lower().replace(" ", "-")
    await safe_goto(page, f"https://www.iimjobs.com/j/{kw}-jobs-1.html")
    for card in await page.query_selector_all(".job-hd, .job-container"):
        try:
            t_el = await card.query_selector("h2 a, .job-title a, a.job-link")
            c_el = await card.query_selector(".company, .job-company, .comp-name")
            d_el = await card.query_selector(".posted, .date-posted, .datelabel")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://www.iimjobs.com" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "IIMJobs", href, age_txt, "India", h))
        except Exception: continue
    print(f"  IIMJobs '{keyword}': {len(jobs)}")
    return jobs


async def scrape_glassdoor(page: Page, keyword: str) -> list[dict]:
    jobs = []
    kw = keyword.replace(" ", "-").lower()
    n  = len(keyword)
    url = f"https://www.glassdoor.co.in/Job/{kw}-jobs-SRCH_KO0,{n}.htm?fromAge=3&sortBy=date_desc"
    await safe_goto(page, url)
    for card in await page.query_selector_all("[data-test='jobListing'], li.react-job-listing"):
        try:
            t_el = await card.query_selector("[data-test='job-link'], a.jobLink")
            c_el = await card.query_selector("[data-test='employer-short-name'], .employer-name")
            d_el = await card.query_selector("[data-test='job-age'], .listing-age, .css-hvni2g")
            l_el = await card.query_selector("[data-test='emp-location'], .location, .css-1v5elnn")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://www.glassdoor.co.in" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            loc     = (await l_el.inner_text()).strip() if l_el else "India"
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Glassdoor", href, age_txt, loc, h))
        except Exception: continue
    print(f"  Glassdoor '{keyword}': {len(jobs)}")
    return jobs


async def scrape_cutshort(page: Page, keyword: str) -> list[dict]:
    """Cutshort.io — popular tech job platform in India"""
    jobs = []
    kw = keyword.replace(" ", "%20")
    await safe_goto(page, f"https://cutshort.io/jobs#!?keywords={kw}&locations=India")
    await page.wait_for_timeout(3500)
    for card in await page.query_selector_all("[class*='JobCard'], .job-card, .job-row"):
        try:
            t_el = await card.query_selector("h2 a, [class*='title'] a, .job-title a")
            c_el = await card.query_selector("[class*='company'], .company-name")
            d_el = await card.query_selector("[class*='date'], time, .posted-date")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://cutshort.io" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Cutshort", href, age_txt, "India", h))
        except Exception: continue
    print(f"  Cutshort '{keyword}': {len(jobs)}")
    return jobs


async def scrape_instahyre(page: Page, keyword: str) -> list[dict]:
    """Instahyre — AI-powered hiring platform, popular in India"""
    jobs = []
    kw = keyword.replace(" ", "%20")
    await safe_goto(page, f"https://www.instahyre.com/search-jobs/?q={kw}&l=India")
    await page.wait_for_timeout(3000)
    for card in await page.query_selector_all(".opportunity-card, [class*='JobCard'], .job-item"):
        try:
            t_el = await card.query_selector("h2 a, .job-title a, [class*='title'] a")
            c_el = await card.query_selector(".company-name, [class*='company']")
            d_el = await card.query_selector(".posted-time, time, [class*='date']")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://www.instahyre.com" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Instahyre", href, age_txt, "India", h))
        except Exception: continue
    print(f"  Instahyre '{keyword}': {len(jobs)}")
    return jobs


async def scrape_freshersworld(page: Page, keyword: str) -> list[dict]:
    """Freshersworld — large India jobs board"""
    jobs = []
    kw = keyword.replace(" ", "+")
    await safe_goto(page, f"https://www.freshersworld.com/jobs/jobdetails/{kw}-jobs-in-India?src=freshersworld")
    for card in await page.query_selector_all(".job-title-name, .jobs-container li"):
        try:
            t_el = await card.query_selector("a.job-title-name, h3 a, .title a")
            c_el = await card.query_selector(".company-name, .comp")
            d_el = await card.query_selector(".posted-date, .date")
            if not t_el: continue
            title   = (await t_el.inner_text()).strip()
            href    = await t_el.get_attribute("href") or ""
            if not href.startswith("http"): href = "https://www.freshersworld.com" + href
            company = (await c_el.inner_text()).strip() if c_el else ""
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, company, "Freshersworld", href, age_txt, "India", h))
        except Exception: continue
    print(f"  Freshersworld '{keyword}': {len(jobs)}")
    return jobs


# ── API scrapers (no browser) ──────────────────────────────────────────────────

async def scrape_greenhouse_api() -> list[dict]:
    """Greenhouse public job board API — no auth needed."""
    import httpx
    jobs = []
    # Companies known to post Salesforce roles and likely use Greenhouse
    companies = [
        "veeva", "medallia", "zuora", "docusign", "ringcentral",
        "zendesk", "mulesoft", "tableau",
        "capgemini", "slalom", "publicissapient",
        "mphasis", "hexaware", "mastek", "persistent",
        "cyient", "coforge", "zs", "syntel",
        "salesforceben", "cloudcoaching", "cloudsolutions",
    ]
    sf_kw = ["salesforce", "lwc", "apex", "crm", "vlocity", "cpq", "einstein"]
    try:
        async with httpx.AsyncClient(timeout=12, follow_redirects=True) as client:
            async def fetch(co: str):
                try:
                    r = await client.get(
                        f"https://boards-api.greenhouse.io/v1/boards/{co}/jobs"
                    )
                    if r.status_code != 200:
                        return
                    for j in r.json().get("jobs", []):
                        t = j.get("title", "")
                        if not any(k in t.lower() for k in sf_kw):
                            continue
                        loc = j.get("location", {}).get("name", "")
                        if loc and "india" not in loc.lower() and "remote" not in loc.lower():
                            continue
                        updated = j.get("updated_at", "")
                        h = parse_age_hours(updated)
                        if not is_recent(h):
                            continue
                        jobs.append(make_job(
                            t,
                            co.replace("-", " ").title(),
                            "Greenhouse",
                            j.get("absolute_url", ""),
                            updated,
                            loc or "India",
                            h,
                        ))
                except Exception:
                    pass
            await asyncio.gather(*[fetch(co) for co in companies])
    except Exception as e:
        print(f"  Greenhouse API error: {e}")
    print(f"  Greenhouse API: {len(jobs)}")
    return jobs


async def scrape_lever_api() -> list[dict]:
    """Lever public posting API — no auth needed."""
    import httpx
    jobs = []
    sf_kw = ["salesforce", "lwc", "apex", "crm", "vlocity", "cpq", "einstein"]
    companies = [
        "salesforce", "slalom", "deloitte", "accenture", "capgemini",
        "cognizant", "publicissapient", "virtusa", "concentrix",
        "wipro", "infosys", "hcl", "mphasis", "persistent",
        "mastek", "zendesk", "veeva", "zuora", "coforge",
    ]
    try:
        async with httpx.AsyncClient(timeout=12, follow_redirects=True) as client:
            async def fetch(co: str):
                try:
                    r = await client.get(f"https://api.lever.co/v0/postings/{co}?mode=json")
                    if r.status_code != 200:
                        return
                    for p in r.json():
                        t = p.get("text", "")
                        if not any(k in t.lower() for k in sf_kw):
                            continue
                        created_ms = p.get("createdAt", 0)
                        if created_ms:
                            h = (datetime.now(timezone.utc).timestamp() - created_ms / 1000) / 3600
                        else:
                            h = 0.0
                        if not is_recent(h):
                            continue
                        loc = p.get("categories", {}).get("location", "India")
                        jobs.append(make_job(
                            t,
                            co.replace("-", " ").title(),
                            "Lever",
                            p.get("hostedUrl", ""),
                            "",
                            loc or "India",
                            h,
                        ))
                except Exception:
                    pass
            await asyncio.gather(*[fetch(co) for co in companies])
    except Exception as e:
        print(f"  Lever API error: {e}")
    print(f"  Lever API: {len(jobs)}")
    return jobs


async def scrape_workable_api() -> list[dict]:
    """Workable public job board API."""
    import httpx
    jobs = []
    sf_kw = ["salesforce", "lwc", "apex", "crm", "cpq"]
    companies = [
        "mastech", "hcltech", "niit-technologies", "infoobjects",
        "cloudmasonry", "7summits", "concentrix-catalyst",
        "cloudcoaching", "simplus", "apexon",
    ]
    try:
        async with httpx.AsyncClient(timeout=10, follow_redirects=True) as client:
            for co in companies:
                try:
                    r = await client.get(
                        f"https://apply.workable.com/api/v1/widget/accounts/{co}/jobs?details=true"
                    )
                    if r.status_code != 200:
                        continue
                    for j in r.json().get("results", []):
                        t = j.get("title", "")
                        if not any(k in t.lower() for k in sf_kw):
                            continue
                        country = (j.get("location") or {}).get("country", "")
                        if country and country.lower() not in ("india", "in", "remote", ""):
                            continue
                        published = j.get("published_on", "")
                        h = parse_age_hours(published)
                        if not is_recent(h):
                            continue
                        jobs.append(make_job(
                            t,
                            j.get("company", co.title()),
                            "Workable",
                            f"https://apply.workable.com/{co}/j/{j.get('shortcode', '')}/",
                            published,
                            country or "India",
                            h,
                        ))
                except Exception:
                    continue
    except Exception as e:
        print(f"  Workable API error: {e}")
    print(f"  Workable API: {len(jobs)}")
    return jobs


async def scrape_ashby_api() -> list[dict]:
    """Ashby HQ public job board API — growing usage among SaaS companies."""
    import httpx
    jobs = []
    sf_kw = ["salesforce", "lwc", "apex", "crm", "cpq", "vlocity"]
    companies = [
        "veeva-systems", "medallia", "ringcentral",
        "zuora", "docusign", "zendesk",
        "salesloft", "outreach", "clari", "gong",
    ]
    try:
        async with httpx.AsyncClient(timeout=10, follow_redirects=True) as client:
            for co in companies:
                try:
                    r = await client.get(
                        "https://jobs.ashbyhq.com/api/non-user-facing/listing/job-board/all-jobs",
                        params={"organizationHostedJobsPageName": co},
                    )
                    if r.status_code != 200:
                        continue
                    for j in (r.json().get("jobPostings") or []):
                        t = j.get("title", "")
                        if not any(k in t.lower() for k in sf_kw):
                            continue
                        locs = j.get("secondaryLocations") or []
                        loc = locs[0].get("name", "India") if locs else "India"
                        published = j.get("publishedDate", "")
                        h = parse_age_hours(published)
                        if not is_recent(h):
                            continue
                        jobs.append(make_job(
                            t, j.get("organizationName", co.title()), "Ashby",
                            f"https://jobs.ashbyhq.com/{co}/{j.get('id', '')}",
                            published, loc, h,
                        ))
                except Exception:
                    continue
    except Exception as e:
        print(f"  Ashby API error: {e}")
    print(f"  Ashby API: {len(jobs)}")
    return jobs


async def scrape_workday_api() -> list[dict]:
    """
    Workday API — only tenants that respond without Cloudflare blocking.
    Most wd1/wd3/wd5 tenants return 422; wd12 (Salesforce) works.
    Uses the country facet ID to filter India results server-side.
    """
    import httpx
    jobs = []
    sf_kw = ["salesforce", "lwc", "apex", "crm", "vlocity", "cpq", "einstein", "mulesoft"]

    # (tenant, wd-number, site-path, display-name, india-country-facet-id)
    # India facet IDs are tenant-specific Workday WID values; discovered per-tenant.
    TENANTS = [
        # Salesforce careers — wd12, India facet verified
        ("salesforce", 12, "External_Career_Site", "Salesforce",
         "CF_-_REC_-_LRV_-_Job_Posting_Anchor_-_Country_from_Job_Posting_Location_Extended",
         "c4f78be1a8f14da0ab49ce1162348a5e"),
    ]

    INDIA_LOCS = {
        "india", "ind", "bengaluru", "bangalore", "hyderabad", "pune", "mumbai",
        "delhi", "noida", "gurugram", "gurgaon", "chennai", "kolkata", "ahmedabad",
        "kochi", "coimbatore", "remote",
    }

    def is_india(loc_text: str) -> bool:
        if not loc_text:
            return True
        lt = loc_text.lower()
        return any(c in lt for c in INDIA_LOCS)

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": UA,
        "X-Requested-With": "XMLHttpRequest",
    }

    async def fetch_tenant(client, tenant, wdver, site, display, facet_param, india_id):
        base = f"https://{tenant}.wd{wdver}.myworkdayjobs.com"
        url  = f"{base}/wday/cxs/{tenant}/{site}/jobs"
        fetched = {}  # ext_path → job dict (dedup by job ID)
        for q in ["salesforce developer", "salesforce admin", "salesforce architect", "salesforce consultant"]:
            body = {
                "searchText": q,
                "limit": 20,
                "offset": 0,
                "appliedFacets": {facet_param: [india_id]},
            }
            try:
                r = await client.post(url, json=body, headers=headers)
                if r.status_code != 200:
                    continue
                for j in r.json().get("jobPostings", []):
                    t = j.get("title", "")
                    if not any(k in t.lower() for k in sf_kw):
                        continue
                    loc_text = j.get("locationsText", "") or ""
                    if not is_india(loc_text):
                        continue
                    posted_txt = j.get("postedOn", "") or ""
                    h = parse_age_hours(posted_txt)
                    if not is_recent(h):
                        continue
                    ext_path = j.get("externalPath", "")
                    job_url = f"{base}{ext_path}" if ext_path else ""
                    key = ext_path or (t + loc_text)
                    if key not in fetched:
                        fetched[key] = make_job(t, display, "Workday", job_url, posted_txt, loc_text or "India", h)
            except Exception:
                pass
        jobs.extend(fetched.values())

    try:
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
            await asyncio.gather(*[
                fetch_tenant(client, t, v, s, d, fp, iid)
                for t, v, s, d, fp, iid in TENANTS
            ])
    except Exception as e:
        print(f"  Workday API error: {e}")
    print(f"  Workday API: {len(jobs)}")
    return jobs


async def scrape_smartrecruiters_api() -> list[dict]:
    """
    SmartRecruiters public API — no auth required.
    Used by IBM, SAP Labs, Oracle, ServiceNow, many SIs.
    GET /v1/companies/{slug}/postings?keyword=salesforce&countryCode=IN
    """
    import httpx
    jobs = []
    sf_kw = ["salesforce", "lwc", "apex", "crm", "cpq", "vlocity", "mulesoft"]

    # Slugs verified from SmartRecruiters job board URLs (company.smartrecruiters.com)
    COMPANIES = [
        ("IBM",                  "IBM"),
        ("SAP",                  "SAP"),
        ("Oracle",               "Oracle"),
        ("ServiceNow",           "ServiceNow"),
        ("Salesforce",           "Salesforce"),
        ("Cognizant",            "Cognizant"),
        ("Capgemini",            "Capgemini"),
        ("Accenture",            "Accenture"),
        ("Deloitte",             "Deloitte"),
        ("Virtusa",              "Virtusa"),
        ("Mphasis",              "Mphasis"),
        ("PersistentSystems",    "Persistent Systems"),
        ("Coforge",              "Coforge"),
        ("Genpact",              "Genpact"),
        ("WNSGlobalServices",    "WNS"),
        ("Zensar",               "Zensar"),
        ("HexawareTechnologies", "Hexaware"),
        ("LTIMindtree",          "LTIMindtree"),
        ("TechMahindraLtd",      "Tech Mahindra"),
        ("Birlasoft",            "Birlasoft"),
        ("Slalom",               "Slalom"),
        ("GlobalLogic",          "GlobalLogic"),
        ("Unison",               "Unison"),
        ("Cyient",               "Cyient"),
    ]

    async def fetch(client, slug, display):
        try:
            r = await client.get(
                f"https://api.smartrecruiters.com/v1/companies/{slug}/postings",
                params={"keyword": "salesforce", "countryCode": "IN", "limit": 50},
            )
            if r.status_code != 200:
                return
            for j in r.json().get("content", []):
                t = j.get("name", "")
                if not any(k in t.lower() for k in sf_kw):
                    continue
                loc = ((j.get("location") or {}).get("city") or
                       (j.get("location") or {}).get("country") or "India")
                rel_date = j.get("releasedDate", "")
                h = parse_age_hours(rel_date)
                if not is_recent(h):
                    continue
                ref = j.get("ref", "")
                url = f"https://jobs.smartrecruiters.com/{slug}/{ref}" if ref else ""
                jobs.append(make_job(t, display, "SmartRecruiters", url, rel_date, loc, h))
        except Exception:
            pass

    try:
        async with httpx.AsyncClient(timeout=12, follow_redirects=True) as client:
            await asyncio.gather(*[fetch(client, s, d) for s, d in COMPANIES])
    except Exception as e:
        print(f"  SmartRecruiters API error: {e}")
    print(f"  SmartRecruiters API: {len(jobs)}")
    return jobs


async def scrape_icims_api() -> list[dict]:
    """
    iCIMS public career portal API — used by many large enterprises.
    GET https://careers-{tenant}.icims.com/jobs/search?ss=1&searchKeyword=salesforce&in_iframe=1&format=json
    """
    import httpx
    jobs = []
    sf_kw = ["salesforce", "lwc", "apex", "crm", "cpq", "vlocity"]

    TENANTS = [
        ("accenture",   "Accenture"),
        ("cognizant",   "Cognizant"),
        ("infosys",     "Infosys"),
        ("wipro",       "Wipro"),
        ("hcl",         "HCL"),
        ("capgemini",   "Capgemini"),
        ("deloitte",    "Deloitte"),
        ("ibm",         "IBM"),
        ("pwc",         "PwC"),
        ("kpmg",        "KPMG"),
        ("ey",          "EY"),
        ("genpact",     "Genpact"),
        ("wns",         "WNS"),
        ("mphasis",     "Mphasis"),
        ("hexaware",    "Hexaware"),
        ("mastech",     "Mastech"),
    ]

    async def fetch(client, tenant, display):
        try:
            r = await client.get(
                f"https://careers-{tenant}.icims.com/jobs/search",
                params={"ss": 1, "searchKeyword": "salesforce", "in_iframe": 1, "format": "json"},
            )
            if r.status_code != 200:
                return
            data = r.json()
            for j in (data.get("jobs") or data.get("searchResults") or []):
                t = j.get("jobtitle") or j.get("title") or ""
                if not any(k in t.lower() for k in sf_kw):
                    continue
                loc = j.get("joblocation") or j.get("location") or "India"
                if "india" not in loc.lower() and "remote" not in loc.lower():
                    continue
                posted = j.get("postdate") or j.get("datePosted") or ""
                h = parse_age_hours(posted)
                if not is_recent(h):
                    continue
                jid = j.get("id") or j.get("jobId") or ""
                url = f"https://careers-{tenant}.icims.com/jobs/{jid}/job" if jid else ""
                jobs.append(make_job(t, display, "iCIMS", url, posted, loc, h))
        except Exception:
            pass

    try:
        async with httpx.AsyncClient(timeout=12, follow_redirects=True) as client:
            await asyncio.gather(*[fetch(client, t, d) for t, d in TENANTS])
    except Exception as e:
        print(f"  iCIMS API error: {e}")
    print(f"  iCIMS API: {len(jobs)}")
    return jobs


async def scrape_tcs_careers(page: Page, keyword: str) -> list[dict]:
    """TCS — own portal, not a standard ATS"""
    jobs = []
    kw = keyword.replace(" ", "%20")
    await safe_goto(
        page,
        f"https://www.tcs.com/careers/tcs-careers-search-jobs#keyword={kw}&country=India",
        timeout=30000,
    )
    await page.wait_for_timeout(4000)
    for card in await page.query_selector_all(".job-card, .job-listing, [class*='job'], tr.jobrow"):
        try:
            t_el = await card.query_selector("a, .job-title, h3, td.jobtitle")
            if not t_el:
                continue
            title = (await t_el.inner_text()).strip()
            if not any(k in title.lower() for k in ["salesforce", "lwc", "apex", "crm", "cpq"]):
                continue
            href = await t_el.get_attribute("href") or ""
            if not href.startswith("http"):
                href = "https://www.tcs.com" + href
            d_el = await card.query_selector(".date, .posted, time, td.date")
            age_txt = (await d_el.inner_text()).strip() if d_el else ""
            h = parse_age_hours(age_txt)
            if is_recent(h):
                jobs.append(make_job(title, "TCS", "TCS Careers", href, age_txt, "India", h))
        except Exception:
            continue
    print(f"  TCS Careers '{keyword}': {len(jobs)}")
    return jobs


async def scrape_salesforce_careers(page: Page, keyword: str) -> list[dict]:
    """Salesforce official careers site"""
    jobs = []
    kw = keyword.replace(" ", "+")
    await safe_goto(
        page,
        f"https://salesforce.wd12.myworkdayjobs.com/External_Career_Site?q={kw}&locationCountry=IND",
        timeout=30000,
    )
    await page.wait_for_timeout(4000)
    for card in await page.query_selector_all("[data-automation-id='jobPostingTitleLink'], .WDWJ a, li[class*='job']"):
        try:
            title = (await card.inner_text()).strip()
            href  = await card.get_attribute("href") or ""
            if not href.startswith("http"):
                href = "https://salesforce.wd12.myworkdayjobs.com" + href
            if title:
                jobs.append(make_job(title, "Salesforce", "Salesforce Careers", href, "Recent", "India", 0.0))
        except Exception:
            continue
    print(f"  Salesforce Careers '{keyword}': {len(jobs)}")
    return jobs


# ── writers ───────────────────────────────────────────────────────────────────

def write_json(jobs: list[dict]):
    seen: set = set()
    unique = []
    for j in jobs:
        key = (j["title"].lower().strip(), (j.get("company") or "").lower().strip())
        if j.get("url") and key not in seen:
            seen.add(key)
            unique.append(j)
    unique.sort(key=lambda j: j.get("age_hours", 0))
    payload = {
        "jobs":       unique,
        "total":      len(unique),
        "scraped_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "platforms":  sorted({j["platform"] for j in unique}),
    }
    JOBS_JSON.write_text(json.dumps(payload, indent=2))


def write_excel(jobs: list[dict]) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salesforce Jobs"
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    headers = ["#", "Job Title", "Company", "Platform", "Posted", "Apply Link"]
    widths  = [5, 45, 30, 15, 15, 80]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    f_even = PatternFill("solid", fgColor="EBF3FB")
    f_odd  = PatternFill("solid", fgColor="FFFFFF")
    seen: set = set()
    unique = []
    for j in jobs:
        key = (j["title"].lower().strip(), (j.get("company") or "").lower().strip())
        if j.get("url") and key not in seen:
            seen.add(key)
            unique.append(j)
    unique.sort(key=lambda j: j.get("age_hours", 0))
    for i, j in enumerate(unique, 1):
        row = i + 1
        fill = f_even if i % 2 == 0 else f_odd
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=j["title"])
        ws.cell(row=row, column=3, value=j.get("company", ""))
        ws.cell(row=row, column=4, value=j["platform"])
        ws.cell(row=row, column=5, value=j.get("posted", ""))
        url = j.get("url", "")
        lc  = ws.cell(row=row, column=6, value=url)
        if url.startswith("http"):
            lc.hyperlink = url
            lc.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[row].height = 15
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    return len(unique)


# ── main ──────────────────────────────────────────────────────────────────────

async def main():
    print(f"\n{'='*60}")
    print("  Salesforce Job Finder — Multi-Platform | India | 72h")
    print(f"{'='*60}\n")

    all_jobs: list[dict] = []

    # ── Phase 1: API scrapers (no browser, run in parallel) ──────────────────
    print("[API scrapers — no browser]")
    api_results = await asyncio.gather(
        scrape_greenhouse_api(),
        scrape_lever_api(),
        scrape_workable_api(),
        scrape_ashby_api(),
        scrape_workday_api(),
        scrape_smartrecruiters_api(),
        scrape_icims_api(),
    )
    for r in api_results:
        all_jobs.extend(r)

    # ── Phase 2: Browser scrapers (one context per scraper per keyword) ───────
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)

        # Company career pages — keyword-independent, run once
        print("\n[Company career pages]")
        async def run_once(fn):
            ctx = await browser.new_context(user_agent=UA)
            pg  = await ctx.new_page()
            try:
                return await fn(pg, "salesforce")
            except Exception as e:
                print(f"  {fn.__name__} error: {e}")
                return []
            finally:
                await ctx.close()

        career_results = await asyncio.gather(
            run_once(scrape_tcs_careers),
            run_once(scrape_salesforce_careers),
        )
        for r in career_results:
            all_jobs.extend(r)

        for keyword in KEYWORDS:
            print(f"\n[{keyword}]")

            async def run(fn, kw=keyword):
                ctx = await browser.new_context(user_agent=UA)
                pg  = await ctx.new_page()
                try:
                    return await fn(pg, kw)
                except Exception as e:
                    print(f"  {fn.__name__} error: {e}")
                    return []
                finally:
                    await ctx.close()

            results = await asyncio.gather(
                run(scrape_linkedin),
                run(scrape_indeed),
                run(scrape_naukri),
                run(scrape_timesjobs),
                run(scrape_foundit),
                run(scrape_shine),
                run(scrape_iimjobs),
                run(scrape_glassdoor),
                run(scrape_cutshort),
                run(scrape_instahyre),
                run(scrape_freshersworld),
            )
            for r in results:
                all_jobs.extend(r)

        await browser.close()

    total = write_excel(all_jobs)
    write_json(all_jobs)
    print(f"\n{'='*60}")
    print(f"  Done! {total} unique jobs → jobs.json + xlsx")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    asyncio.run(main())
